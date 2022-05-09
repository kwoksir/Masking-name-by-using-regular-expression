import openpyxl
import re

wb = openpyxl.load_workbook('demo.xlsx')
sheet = wb.worksheets[0]

for i in range(sheet.max_row):
    sheet.cell(row=i+1, column=2).value = re.sub(r'[a-z]', '*', sheet.cell(row=i+1,column=1).value)
 
wb.save("demo.xlsx") 
